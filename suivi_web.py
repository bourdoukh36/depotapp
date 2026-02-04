import os
import streamlit as st
import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
import pandas as pd

# ================= CONFIG GOOGLE SHEETS =================
SCOPE = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive"
]

creds = ServiceAccountCredentials.from_json_keyfile_dict(
    st.secrets["gcp_service_account"],
    SCOPE
)
client = gspread.authorize(creds)
SHEET_NAME = "suivi des opérations"

# ================= DONNÉES FIXES =================
SERRES = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
DELTAS = [str(i) for i in range(1, 33)]
CULTURES = ['tomate', 'pastèque', 'poivron', 'concombre', 'laitue', 'ciboulette', 'courgette', 'herbes aromatiques']
TRAITEMENTS = ['fongicide', 'insecticide', 'acaricide', 'insecticide/acaricide', 'raticide', 'bio-stimulant',
               'désinfectant', 'engrais foliaire']
TRAITEMENT_COLORS = {
    'fongicide': '#ff9999', 'insecticide': '#99ff99', 'acaricide': '#99ccff',
    'insecticide/acaricide': '#ffcc99', 'raticide': '#cccccc', 'bio-stimulant': '#ffff99',
    'désinfectant': '#ffccff', 'engrais foliaire': '#ccffcc'
}
SOLUTIONS_IRRI = ['AB', 'CD', 'M', 'Urée', 'enracineur', 'désinfectant']
ECS = ['1.6', '1.8', '2', '2.5', '3', '3.5', '4']

EXCEL_PRODUITS = "produits.xlsx"
EXCEL_OPERATIONS = "operations.xlsx"

# ================= CREATION AUTOMATIQUE EXCEL =================
if not os.path.exists(EXCEL_PRODUITS):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Designation", "Dose", "Cible"])
    wb.save(EXCEL_PRODUITS)

if not os.path.exists(EXCEL_OPERATIONS):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Serre", "Delta", "Culture", "Traitement", "Solution", "ECS", "Remarques"])
    wb.save(EXCEL_OPERATIONS)

# ================= FONCTIONS =================
def charger_produits():
    wb = openpyxl.load_workbook(EXCEL_PRODUITS)
    ws = wb.active
    return [row for row in ws.iter_rows(min_row=2, values_only=True) if row]

def ajouter_produit(designation, dose, cible):
    wb = openpyxl.load_workbook(EXCEL_PRODUITS)
    ws = wb.active
    ws.append([designation, dose, cible])
    wb.save(EXCEL_PRODUITS)

def charger_operations():
    wb = openpyxl.load_workbook(EXCEL_OPERATIONS)
    ws = wb.active
    return [row for row in ws.iter_rows(min_row=2, values_only=True) if row]

def ajouter_operation(date, serre, delta, culture, traitement, solution, ecs, remarques):
    wb = openpyxl.load_workbook(EXCEL_OPERATIONS)
    ws = wb.active
    ws.append([date, serre, delta, culture, traitement, solution, ecs, remarques])
    wb.save(EXCEL_OPERATIONS)

def filter_operations(ops, serre_filter, culture_filter, traitement_filter):
    filtered = []
    for op in ops:
        _, serre, _, culture, traitement, _, _, _ = op
        if (serre_filter != "Toutes" and serre != serre_filter):
            continue
        if (culture_filter != "Toutes" and culture != culture_filter):
            continue
        if (traitement_filter != "Tous" and traitement != traitement_filter):
            continue
        filtered.append(op)
    return filtered

# ================= STYLE CSS MOBILE =================
st.markdown("""
<style>
body {font-family: sans-serif;}
.card {padding: 10px; margin: 5px 0; border-radius: 8px; box-shadow: 1px 1px 4px #aaa;}
.card-title {font-weight: bold; font-size: 16px; margin-bottom: 3px;}
.card-content {font-size: 14px;}
.stButton>button {background-color:#0080ff; color:white; width:100%; margin-top:5px;}
@media (max-width: 600px){
    .card-title {font-size:14px;}
    .card-content {font-size:12px;}
}
</style>
""", unsafe_allow_html=True)

# ================= INTERFACE =================
st.title("Suivi des opérations pépinière 🌱")

# ---------- PRODUITS ----------
st.subheader("Produits disponibles")
produits = charger_produits()
if produits:
    for p in produits:
        st.markdown(
            f"<div class='card'><div class='card-title'>{p[0]}</div><div class='card-content'><b>Dose:</b> {p[1]} | <b>Cible:</b> {p[2]}</div></div>",
            unsafe_allow_html=True
        )

st.markdown("### Ajouter un produit")
new_designation = st.text_input("Désignation", key="prod_name")
new_dose = st.text_input("Dose", key="prod_dose")
new_cible = st.text_input("Cible", key="prod_cible")
if st.button("Ajouter produit"):
    if new_designation and new_dose and new_cible:
        ajouter_produit(new_designation, new_dose, new_cible)
        st.success(f"Produit '{new_designation}' ajouté !")
    else:
        st.warning("Veuillez remplir tous les champs.")

# ---------- OPERATIONS ----------
st.markdown("---")
st.subheader("Ajouter une opération")
op_serre = st.selectbox("Serre", ["Toutes"] + SERRES)
op_delta = st.selectbox("Delta", ["Toutes"] + DELTAS)
op_culture = st.selectbox("Culture", ["Toutes"] + CULTURES)
op_traitement = st.selectbox("Traitement", ["Tous"] + TRAITEMENTS)
op_solution = st.selectbox("Solution", SOLUTIONS_IRRI)
op_ecs = st.selectbox("ECS", ECS)
op_remarques = st.text_input("Remarques")
if st.button("Ajouter opération"):
    date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    ajouter_operation(date_str, op_serre, op_delta, op_culture, op_traitement, op_solution, op_ecs, op_remarques)
    st.success(f"Opération ajoutée pour {op_culture} dans serre {op_serre}")

# ---------- AFFICHAGE OPERATIONS ----------
st.markdown("---")
st.subheader("Liste des opérations")
ops = charger_operations()
ops_filtered = filter_operations(ops, op_serre, op_culture, op_traitement)

# Pagination
items_per_page = 8
page = st.number_input("Page", min_value=1, max_value=(len(ops_filtered)-1)//items_per_page+1, value=1)
start = (page-1)*items_per_page
end = start + items_per_page

for op in ops_filtered[start:end]:
    date, serre, delta, culture, traitement, solution, ecs, remarques = op
    color = TRAITEMENT_COLORS.get(traitement, "#f0f8ff")
    with st.expander(f"{culture} - Serre {serre} Delta {delta} | {traitement}"):
        st.markdown(
            f"<div class='card' style='background-color:{color};'><div class='card-content'><b>Date:</b> {date} | <b>Solution:</b> {solution} | <b>ECS:</b> {ecs}<br><b>Remarques:</b> {remarques}</div></div>",
            unsafe_allow_html=True
        )

# ---------- EXPORT EXCEL ----------
if st.button("Exporter les opérations filtrées en Excel"):
    df = pd.DataFrame(ops_filtered, columns=["Date", "Serre", "Delta", "Culture", "Traitement", "Solution", "ECS", "Remarques"])
    export_file = f"operations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    df.to_excel(export_file, index=False)
    with open(export_file, "rb") as f:
        st.download_button("Télécharger Excel", f, file_name=export_file)
